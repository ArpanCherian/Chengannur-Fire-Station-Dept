from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
import csv
from django.http import HttpResponse
from datetime import datetime
from django.contrib import messages
from .models import userdetails
from .models import admindetails
from .models import firewater
from .models import generalIncident
from .models import AssistanceCall  

def home(request):
    return render(request, 'index.html')

def about(request):
    return render(request, 'about.html')

def contact(request):
    return render(request, 'contact.html')



# User login form view
def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            user = userdetails.objects.get(username=username)
            if password == user.password:
                # Login success logic e.g., create session
                request.session['username'] = username  # simple session example
                messages.success(request, f"You have successfully logged in!")
                return redirect('userdashboard')
            else:
                messages.error(request, "Invalid password")
        except userdetails.DoesNotExist:
            messages.error(request, "Username not found")

    return render(request, 'user_login.html')

def logout(request):
    request.session.flush()
    return redirect('index')




def admin_login(request):
    if request.method == 'POST':
        username = request.POST.get('adminusername')
        password = request.POST.get('adminpassword')

        try:
            admin = admindetails.objects.get(adminusername=username)
            if password == admin.adminpassword:
                # Admin login success logic e.g., create session
                request.session['username'] = username  # simple session example
                messages.success(request, f"You have successfully logged in!")
                return redirect('admindashboard')
            else:
                messages.error(request, "Invalid password")
        except admindetails.DoesNotExist:
            messages.error(request, "Admin username not found")
    # admin login logic here or render admin login template
    return render(request, 'admin_login.html')




def userdashboard(request):
    if 'username' not in request.session:
        return redirect('user_login')

    username = request.session['username']

    # Fetch all user reports
    firewater_cases = list(firewater.objects.filter(userName=username).order_by('-form_date'))
    general_cases = list(generalIncident.objects.filter(userName=username).order_by('-form_date'))
    assistant_cases = list(AssistanceCall.objects.filter(userName=username).order_by('-form_date'))

    # Add model_type attribute to distinguish cases
    for case in firewater_cases:
        case.model_type = 'firewater'
    for case in general_cases:
        case.model_type = 'general'
    for case in assistant_cases:
        case.model_type = 'assistance'

    # Combine all into one list
    all_cases = firewater_cases + general_cases + assistant_cases
    all_cases_sorted = sorted(all_cases, key=lambda case: case.form_date, reverse=True)

    # Stats calculations
    total_reports = len(all_cases)
    fire_cases_count = len([c for c in firewater_cases if c.incident == 'Fire'])
    water_cases_count = len([c for c in firewater_cases if c.incident == 'Water'])
    other_cases_count = len(general_cases) + len(assistant_cases)

    context = {
        'recent_cases': all_cases_sorted[:10],
        'username': username,
        'total_reports': total_reports,
        'fire_cases': fire_cases_count,
        'water_cases': water_cases_count,
        'other_cases': other_cases_count,
    }
    return render(request, 'userdashboard.html', context)




def reportcase(request):
    return render(request, 'reportcase.html')




def admindashboard(request):
    if 'username' not in request.session:
        return redirect('admin_login')

    username = request.session['username']

    # Fetch all user reports
    firewater_cases = list(firewater.objects.filter(userName=username).order_by('-form_date'))
    general_cases = list(generalIncident.objects.filter(userName=username).order_by('-form_date'))
    assistant_cases = list(AssistanceCall.objects.filter(userName=username).order_by('-form_date'))

    # Add model_type attribute to distinguish cases
    for case in firewater_cases:
        case.model_type = 'firewater'
    for case in general_cases:
        case.model_type = 'general'
    for case in assistant_cases:
        case.model_type = 'assistance'

    # Combine all into one list
    all_cases = firewater_cases + general_cases + assistant_cases
    all_cases_sorted = sorted(all_cases, key=lambda case: case.form_date, reverse=True)

    # Stats calculations
    total_reports = len(all_cases)
    fire_cases_count = len([c for c in firewater_cases if c.incident == 'Fire'])
    water_cases_count = len([c for c in firewater_cases if c.incident == 'Water'])
    other_cases_count = len(general_cases) + len(assistant_cases)

    context = {
        'recent_cases': all_cases_sorted[:10],
        'username': username,
        'total_reports': total_reports,
        'fire_cases': fire_cases_count,
        'water_cases': water_cases_count,
        'other_cases': other_cases_count,
    }
    return render(request, 'admindashboard.html', context)






def adminreportcase(request):
    return render(request, 'adminreportcase.html')


def fireform(request):
    return render(request, 'fireform.html')

def adminfireform(request):
    return render(request, 'adminfireform.html')

def waterform(request):
    return render(request, 'waterform.html')

def adminwaterform(request):
    return render(request, 'adminwaterform.html')

def generalincidentform(request):
    return render(request, 'generalincidentform.html')

def admingeneralincidentform(request):
    return render(request, 'admingeneralincidentform.html')

def assistcalls(request):
    return render(request, 'assistcalls.html')

def adminassistcalls(request):
    return render(request, 'adminassistcalls.html')



def firewater_report(request):
    if request.method == 'POST':
        data = request.POST
        obj = firewater(
            userName = data.get('userName'),
            form_date = data.get('form_date'),
            incident = data.get('incident'),
            call_number = data.get('call_number'),
            call_received = data.get('call_received'),
            time_left_station = data.get('time_left_station'),
            time_reached_scene = data.get('time_reached_scene'),
            time_returned = data.get('time_returned'),
            occupancy_type = data.get('occupancy_type'),
            construction_type = data.get('construction_type'),
            owner_details = data.get('owner_details'),
            electrical_chemicals = data.get('electrical_chemicals'),
            hazardous_materials = data.get('hazardous_materials'),
            weather_conditions = data.get('weather_conditions'),
            caller_contact = data.get('caller_contact'),
            source_of_call = data.get('source_of_call'),
            nature_incident = data.get('nature_incident'),
            exact_location = data.get('exact_location'),
            owner_occupant = data.get('owner_occupant'),
            premises_occupancy = data.get('premises_occupancy'),
            building_details = data.get('building_details'),
            premises_chemicals = data.get('premises_chemicals'),
            premises_hazardous = data.get('premises_hazardous'),
            deaths_male = data.get('deaths_male'),
            deaths_female = data.get('deaths_female'),
            deaths_adult = data.get('deaths_adult'),
            deaths_child = data.get('deaths_child'),
            injuries_male = data.get('injuries_male'),
            injuries_female = data.get('injuries_female'),
            injuries_adult = data.get('injuries_adult'),
            injuries_child = data.get('injuries_child'),
            rescued_male = data.get('rescued_male'),
            rescued_female = data.get('rescued_female'),
            rescued_adult = data.get('rescued_adult'),
            rescued_child = data.get('rescued_child'),
            fire_personnel_injuries = data.get('fire_personnel_injuries'),
            animals_rescued = data.get('animals_rescued'),
            animals_lost = data.get('animals_lost'),
            hospital_doctor = data.get('hospital_doctor'),
            ambulance_notified = data.get('ambulance_notified'),
            appliances_crews = data.get('appliances_crews'),
            officer_in_charge = data.get('officer_in_charge'),
            equipment_used = data.get('equipment_used'),
            property_removed = data.get('property_removed'),
            property_saved = data.get('property_saved'),
            property_lost = data.get('property_lost'),
            building_damage = data.get('building_damage'),
            items_destroyed = data.get('items_destroyed'),
            major_loss_cause = data.get('major_loss_cause')
        )
        obj.save()
        messages.success(request, 'Form submitted successfully!')
        return redirect('reportcase')  # Replace with your success url name
    messages.error(request, 'Please correct the errors below.')
    return render(request, 'reportcase.html')  # Render your form template

def adminfirewater_report(request):
    if request.method == 'POST':
        data = request.POST
        obj = firewater(
            userName = data.get('userName'),
            form_date = data.get('form_date'),
            incident = data.get('incident'),
            call_number = data.get('call_number'),
            call_received = data.get('call_received'),
            time_left_station = data.get('time_left_station'),
            time_reached_scene = data.get('time_reached_scene'),
            time_returned = data.get('time_returned'),
            occupancy_type = data.get('occupancy_type'),
            construction_type = data.get('construction_type'),
            owner_details = data.get('owner_details'),
            electrical_chemicals = data.get('electrical_chemicals'),
            hazardous_materials = data.get('hazardous_materials'),
            weather_conditions = data.get('weather_conditions'),
            caller_contact = data.get('caller_contact'),
            source_of_call = data.get('source_of_call'),
            nature_incident = data.get('nature_incident'),
            exact_location = data.get('exact_location'),
            owner_occupant = data.get('owner_occupant'),
            premises_occupancy = data.get('premises_occupancy'),
            building_details = data.get('building_details'),
            premises_chemicals = data.get('premises_chemicals'),
            premises_hazardous = data.get('premises_hazardous'),
            deaths_male = data.get('deaths_male'),
            deaths_female = data.get('deaths_female'),
            deaths_adult = data.get('deaths_adult'),
            deaths_child = data.get('deaths_child'),
            injuries_male = data.get('injuries_male'),
            injuries_female = data.get('injuries_female'),
            injuries_adult = data.get('injuries_adult'),
            injuries_child = data.get('injuries_child'),
            rescued_male = data.get('rescued_male'),
            rescued_female = data.get('rescued_female'),
            rescued_adult = data.get('rescued_adult'),
            rescued_child = data.get('rescued_child'),
            fire_personnel_injuries = data.get('fire_personnel_injuries'),
            animals_rescued = data.get('animals_rescued'),
            animals_lost = data.get('animals_lost'),
            hospital_doctor = data.get('hospital_doctor'),
            ambulance_notified = data.get('ambulance_notified'),
            appliances_crews = data.get('appliances_crews'),
            officer_in_charge = data.get('officer_in_charge'),
            equipment_used = data.get('equipment_used'),
            property_removed = data.get('property_removed'),
            property_saved = data.get('property_saved'),
            property_lost = data.get('property_lost'),
            building_damage = data.get('building_damage'),
            items_destroyed = data.get('items_destroyed'),
            major_loss_cause = data.get('major_loss_cause')
        )
        obj.save()
        messages.success(request, 'Form submitted successfully!')
        return redirect('adminreportcase')  # Replace with your success url name
    messages.error(request, 'Please correct the errors below.')
    return render(request, 'adminreportcase.html')  # Render your form template


def generalincident_report(request):
    if request.method == 'POST':
        data = request.POST
        obj = generalIncident(
            userName = data.get('userName'),
            form_date = data.get('form_date'),
            incident = data.get('incident'),
            call_number = data.get('call_number'),
            call_received = data.get('call_received'),
            time_left_station = data.get('time_left_station'),
            time_reached_scene = data.get('time_reached_scene'),
            time_returned = data.get('time_returned'),
            occupancy_type = data.get('occupancy_type'),
            construction_type = data.get('construction_type'),
            owner_details = data.get('owner_details'),
            electrical_chemicals = data.get('electrical_chemicals'),
            hazardous_materials = data.get('hazardous_materials'),
            weather_conditions = data.get('weather_conditions'),
            caller_contact = data.get('caller_contact'),
            source_of_call = data.get('source_of_call'),
            nature_incident = data.get('nature_incident'),
            exact_location = data.get('exact_location'),
            owner_occupant = data.get('owner_occupant'),
            premises_occupancy = data.get('premises_occupancy'),
            building_details = data.get('building_details'),
            premises_chemicals = data.get('premises_chemicals'),
            premises_hazardous = data.get('premises_hazardous'),
            deaths_male = data.get('deaths_male'),
            deaths_female = data.get('deaths_female'),
            deaths_adult = data.get('deaths_adult'),
            deaths_child = data.get('deaths_child'),
            injured = data.get('injured'),
            rescued_male = data.get('rescued_male'),
            rescued_female = data.get('rescued_female'),
            rescued_adult = data.get('rescued_adult'),
            rescued_child = data.get('rescued_child'),
            fire_personnel_injuries = data.get('fire_personnel_injuries'),
            animals_rescued = data.get('animals_rescued'),
            animals_lost = data.get('animals_lost'),
            hospital_doctor = data.get('hospital_doctor'),
            ambulance_notified = data.get('ambulance_notified'),
            appliances_crews = data.get('appliances_crews'),
            officer_in_charge = data.get('officer_in_charge'),
            equipment_used = data.get('equipment_used'),
            property_removed = data.get('property_removed'),
            property_saved = data.get('property_saved'),
            property_lost = data.get('property_lost'),
            building_damage = data.get('building_damage'),
            items_destroyed = data.get('items_destroyed'),
            major_loss_cause = data.get('major_loss_cause')
        )
        obj.save()
        messages.success(request, 'Form submitted successfully!')
        return redirect('reportcase')  # Replace with your success url name
    messages.error(request, 'Please correct the errors below.')
    return render(request, 'generalincidentform.html') 


def admingeneralincident_report(request):
    if request.method == 'POST':
        data = request.POST
        obj = generalIncident(
            userName = data.get('userName'),
            form_date = data.get('form_date'),
            incident = data.get('incident'),
            call_number = data.get('call_number'),
            call_received = data.get('call_received'),
            time_left_station = data.get('time_left_station'),
            time_reached_scene = data.get('time_reached_scene'),
            time_returned = data.get('time_returned'),
            occupancy_type = data.get('occupancy_type'),
            construction_type = data.get('construction_type'),
            owner_details = data.get('owner_details'),
            electrical_chemicals = data.get('electrical_chemicals'),
            hazardous_materials = data.get('hazardous_materials'),
            weather_conditions = data.get('weather_conditions'),
            caller_contact = data.get('caller_contact'),
            source_of_call = data.get('source_of_call'),
            nature_incident = data.get('nature_incident'),
            exact_location = data.get('exact_location'),
            owner_occupant = data.get('owner_occupant'),
            premises_occupancy = data.get('premises_occupancy'),
            building_details = data.get('building_details'),
            premises_chemicals = data.get('premises_chemicals'),
            premises_hazardous = data.get('premises_hazardous'),
            deaths_male = data.get('deaths_male'),
            deaths_female = data.get('deaths_female'),
            deaths_adult = data.get('deaths_adult'),
            deaths_child = data.get('deaths_child'),
            injured = data.get('injured'),
            rescued_male = data.get('rescued_male'),
            rescued_female = data.get('rescued_female'),
            rescued_adult = data.get('rescued_adult'),
            rescued_child = data.get('rescued_child'),
            fire_personnel_injuries = data.get('fire_personnel_injuries'),
            animals_rescued = data.get('animals_rescued'),
            animals_lost = data.get('animals_lost'),
            hospital_doctor = data.get('hospital_doctor'),
            ambulance_notified = data.get('ambulance_notified'),
            appliances_crews = data.get('appliances_crews'),
            officer_in_charge = data.get('officer_in_charge'),
            equipment_used = data.get('equipment_used'),
            property_removed = data.get('property_removed'),
            property_saved = data.get('property_saved'),
            property_lost = data.get('property_lost'),
            building_damage = data.get('building_damage'),
            items_destroyed = data.get('items_destroyed'),
            major_loss_cause = data.get('major_loss_cause')
        )
        obj.save()
        messages.success(request, 'Form submitted successfully!')
        return redirect('adminreportcase')  # Replace with your success url name
    messages.error(request, 'Please correct the errors below.')
    return render(request, 'admingeneralincidentform.html') 



def assistancecall_report(request):
    if request.method == 'POST':
        data = request.POST
        obj = AssistanceCall(
            userName = data.get('userName'),
            form_date = data.get('form_date'),
            incident_type = data.get('incident_type'),
            call_number = data.get('call_number'),
            call_received = data.get('call_received'),
            time_left_station = data.get('time_left_station'),
            time_reached_scene = data.get('time_reached_scene'),
            time_returned = data.get('time_returned'),
            assistance_details = data.get('assistance_details'),
            electrical_chemicals = data.get('electrical_chemicals'),
            hazardous_materials = data.get('hazardous_materials'),
            weather_conditions = data.get('weather_conditions'),
            caller_contact = data.get('caller_contact'),
            source_of_call = data.get('source_of_call'),
            exact_location = data.get('exact_location'),
        )
        obj.save()
        messages.success(request, 'Form submitted successfully!')
        return redirect('reportcase')  # Replace with your success url name
    messages.error(request, 'Please correct the errors below.')
    return render(request, 'assistcalls.html') 



def adminassistancecall_report(request):
    if request.method == 'POST':
        data = request.POST
        obj = AssistanceCall(
            userName = data.get('userName'),
            form_date = data.get('form_date'),
            incident_type = data.get('incident_type'),
            call_number = data.get('call_number'),
            call_received = data.get('call_received'),
            time_left_station = data.get('time_left_station'),
            time_reached_scene = data.get('time_reached_scene'),
            time_returned = data.get('time_returned'),
            assistance_details = data.get('assistance_details'),
            electrical_chemicals = data.get('electrical_chemicals'),
            hazardous_materials = data.get('hazardous_materials'),
            weather_conditions = data.get('weather_conditions'),
            caller_contact = data.get('caller_contact'),
            source_of_call = data.get('source_of_call'),
            exact_location = data.get('exact_location'),
        )
        obj.save()
        messages.success(request, 'Form submitted successfully!')
        return redirect('adminreportcase')  # Replace with your success url name
    messages.error(request, 'Please correct the errors below.')
    return render(request, 'adminassistcalls.html') 



def case_detail(request, model_type, case_id):
    if 'username' not in request.session:
        return redirect('user_login')
    
    username = request.session['username']
    
    # Fetch the case based on model type
    if model_type == 'firewater':
        case = get_object_or_404(firewater, id=case_id, userName=username)
    elif model_type == 'general':
        case = get_object_or_404(generalIncident, id=case_id, userName=username)
    elif model_type == 'assistance':
        case = get_object_or_404(AssistanceCall, id=case_id, userName=username)
    else:
        return redirect('userdashboard')
    
    context = {
        'case': case,
        'model_type': model_type,
        'username': username,
    }
    return render(request, 'case_detail.html', context)



def admincase_detail(request, model_type, case_id):
    if 'username' not in request.session:
        return redirect('admin_login')
    
    username = request.session['username']
    
    # Fetch the case based on model type
    if model_type == 'firewater':
        case = get_object_or_404(firewater, id=case_id, userName=username)
    elif model_type == 'general':
        case = get_object_or_404(generalIncident, id=case_id, userName=username)
    elif model_type == 'assistance':
        case = get_object_or_404(AssistanceCall, id=case_id, userName=username)
    else:
        return redirect('admindashboard')
    
    context = {
        'case': case,
        'model_type': model_type,
        'username': username,
    }
    return render(request, 'admincase_detail.html', context)



def adminviewreport(request):
    qs_fire = list(firewater.objects.all())
    qs_general = list(generalIncident.objects.all())
    qs_assist = list(AssistanceCall.objects.all())

    month = request.GET.get('month')
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    
    if month:
        y, m = month.split('-')
        qs_fire = [c for c in qs_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_general = [c for c in qs_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_assist = [c for c in qs_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    if start:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date >= start_date]
        qs_general = [c for c in qs_general if c.form_date >= start_date]
        qs_assist = [c for c in qs_assist if c.form_date >= start_date]
    
    if end:
        end_date = datetime.strptime(end, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date <= end_date]
        qs_general = [c for c in qs_general if c.form_date <= end_date]
        qs_assist = [c for c in qs_assist if c.form_date <= end_date]

    # Add model_type to each case
    for case in qs_fire:
        case.model_type = 'firewater'
    for case in qs_general:
        case.model_type = 'general'
    for case in qs_assist:
        case.model_type = 'assistance'

    all_reports = qs_fire + qs_general + qs_assist
    all_reports = sorted(all_reports, key=lambda c: c.form_date, reverse=True)

    return render(request, 'adminviewreport.html', {'all_reports': all_reports})



def clean_time(value):
    if not value:
        return value
    # If value looks like a list/tuple, take first element
    if isinstance(value, (list, tuple)):
        value = value[0]
    value = str(value).strip().replace('"','').replace("'",'')
    try:
        return datetime.strptime(value, "%I:%M %p").strftime("%H:%M")
    except Exception:
        try:
            return datetime.strptime(value, "%H:%M").strftime("%H:%M")
        except Exception:
            try:
                return datetime.strptime(value, "%H:%M:%S").strftime("%H:%M")
            except Exception:
                return ""  # fallback for invalid

def edit_case(request, model_type, case_id):
    if 'username' not in request.session:
        return redirect('admin_login')
    if model_type == 'firewater':
        case = get_object_or_404(firewater, id=case_id)
        template = 'edit_firewater.html'
    elif model_type == 'general':
        case = get_object_or_404(generalIncident, id=case_id)
        template = 'edit_general.html'
    elif model_type == 'assistance':
        case = get_object_or_404(AssistanceCall, id=case_id)
        template = 'edit_assistance.html'
    else:
        return redirect('adminviewreport')
    time_fields = ["call_received", "time_left_station", "time_reached_scene", "time_returned","ambulance_notified"]
    if request.method == 'POST':
        # Debug: print for investigation
        print("POST DATA:", dict(request.POST))
        for field, value in request.POST.items():
            if hasattr(case, field) and field != 'csrfmiddlewaretoken':
                if field in time_fields:
                    value = clean_time(value)
                setattr(case, field, value)
        case.save()
        messages.success(request, 'Report updated successfully!')
        return redirect('adminviewreport')
    context = {
        'case': case,
        'model_type': model_type,
    }
    return render(request, template, context)






def delete_case(request, model_type, case_id):
    if 'username' not in request.session:
        return redirect('admin_login')
    
    # Fetch and delete the case based on model type
    if model_type == 'firewater':
        case = get_object_or_404(firewater, id=case_id)
    elif model_type == 'general':
        case = get_object_or_404(generalIncident, id=case_id)
    elif model_type == 'assistance':
        case = get_object_or_404(AssistanceCall, id=case_id)
    else:
        return redirect('adminviewreport')
    
    case.delete()
    messages.success(request, 'Report deleted successfully!')
    return redirect('adminviewreport')




from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import redirect

def download_reports(request):
    if 'username' not in request.session:
        return redirect('admin_login')
    
    # Get filtered reports
    qs_fire = list(firewater.objects.all())
    qs_general = list(generalIncident.objects.all())
    qs_assist = list(AssistanceCall.objects.all())

    month = request.GET.get('month')
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    
    # Apply filters
    if month:
        y, m = month.split('-')
        qs_fire = [c for c in qs_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_general = [c for c in qs_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_assist = [c for c in qs_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    if start:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date >= start_date]
        qs_general = [c for c in qs_general if c.form_date >= start_date]
        qs_assist = [c for c in qs_assist if c.form_date >= start_date]
    
    if end:
        end_date = datetime.strptime(end, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date <= end_date]
        qs_general = [c for c in qs_general if c.form_date <= end_date]
        qs_assist = [c for c in qs_assist if c.form_date <= end_date]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Incident Reports"
    
    # Define styles
    title_font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    
    section_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    label_font = Font(name='Arial', size=10, bold=True)
    label_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Row counter
    current_row = 1
    
    # Main Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = 'FIRE & RESCUE SERVICES - INCIDENT REPORTS'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    ws.row_dimensions[current_row].height = 35
    current_row += 1
    
    # Report Details
    ws.merge_cells(f'A{current_row}:C{current_row}')
    report_info = ws[f'A{current_row}']
    report_info.value = f'Report Generated: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
    report_info.font = Font(name='Arial', size=10, italic=True)
    report_info.alignment = left_alignment
    report_info.border = thin_border
    
    if month:
        ws.merge_cells(f'D{current_row}:F{current_row}')
        period_cell = ws[f'D{current_row}']
        period_cell.value = f'Period: {month}'
        period_cell.font = Font(name='Arial', size=10, italic=True)
        period_cell.alignment = Alignment(horizontal='right', vertical='center')
        period_cell.border = thin_border
    elif start and end:
        ws.merge_cells(f'D{current_row}:F{current_row}')
        period_cell = ws[f'D{current_row}']
        period_cell.value = f'Period: {start} to {end}'
        period_cell.font = Font(name='Arial', size=10, italic=True)
        period_cell.alignment = Alignment(horizontal='right', vertical='center')
        period_cell.border = thin_border
    else:
        ws.merge_cells(f'D{current_row}:F{current_row}')
        period_cell = ws[f'D{current_row}']
        period_cell.value = 'Period: All Records'
        period_cell.font = Font(name='Arial', size=10, italic=True)
        period_cell.alignment = Alignment(horizontal='right', vertical='center')
        period_cell.border = thin_border
    
    ws.row_dimensions[current_row].height = 20
    current_row += 2
    
    # Prepare all data
    all_reports = []
    
    # Process firewater reports
    for obj in qs_fire:
        all_reports.append({
            'type': 'Firewater',
            'date': obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            'sort_date': obj.form_date if obj.form_date else datetime.min.date(),
            'data': {
                'userName': obj.userName,
                'incident': getattr(obj, 'incident', 'N/A'),
                'call_number': getattr(obj, 'call_number', 'N/A'),
                'call_received': getattr(obj, 'call_received', 'N/A'),
                'time_left_station': getattr(obj, 'time_left_station', 'N/A'),
                'time_reached_scene': getattr(obj, 'time_reached_scene', 'N/A'),
                'time_returned': getattr(obj, 'time_returned', 'N/A'),
                'occupancy_type': getattr(obj, 'occupancy_type', 'N/A'),
                'construction_type': getattr(obj, 'construction_type', 'N/A'),
                'owner_details': getattr(obj, 'owner_details', 'N/A'),
                'electrical_chemicals': getattr(obj, 'electrical_chemicals', 'N/A'),
                'hazardous_materials': getattr(obj, 'hazardous_materials', 'N/A'),
                'weather_conditions': getattr(obj, 'weather_conditions', 'N/A'),
                'caller_contact': getattr(obj, 'caller_contact', 'N/A'),
                'source_of_call': getattr(obj, 'source_of_call', 'N/A'),
                'nature_incident': getattr(obj, 'nature_incident', 'N/A'),
                'exact_location': getattr(obj, 'exact_location', 'N/A'),
                'owner_occupant': getattr(obj, 'owner_occupant', 'N/A'),
                'premises_occupancy': getattr(obj, 'premises_occupancy', 'N/A'),
                'building_details': getattr(obj, 'building_details', 'N/A'),
                'premises_chemicals': getattr(obj, 'premises_chemicals', 'N/A'),
                'premises_hazardous': getattr(obj, 'premises_hazardous', 'N/A'),
                'deaths_male': getattr(obj, 'deaths_male', 'N/A'),
                'deaths_female': getattr(obj, 'deaths_female', 'N/A'),
                'deaths_adult': getattr(obj, 'deaths_adult', 'N/A'),
                'deaths_child': getattr(obj, 'deaths_child', 'N/A'),
                'injuries_male': getattr(obj, 'injuries_male', 'N/A'),
                'injuries_female': getattr(obj, 'injuries_female', 'N/A'),
                'injuries_adult': getattr(obj, 'injuries_adult', 'N/A'),
                'injuries_child': getattr(obj, 'injuries_child', 'N/A'),
                'rescued_male': getattr(obj, 'rescued_male', 'N/A'),
                'rescued_female': getattr(obj, 'rescued_female', 'N/A'),
                'rescued_adult': getattr(obj, 'rescued_adult', 'N/A'),
                'rescued_child': getattr(obj, 'rescued_child', 'N/A'),
                'fire_personnel_injuries': getattr(obj, 'fire_personnel_injuries', 'N/A'),
                'animals_rescued': getattr(obj, 'animals_rescued', 'N/A'),
                'animals_lost': getattr(obj, 'animals_lost', 'N/A'),
                'hospital_doctor': getattr(obj, 'hospital_doctor', 'N/A'),
                'ambulance_notified': getattr(obj, 'ambulance_notified', 'N/A'),
                'appliances_crews': getattr(obj, 'appliances_crews', 'N/A'),
                'officer_in_charge': getattr(obj, 'officer_in_charge', 'N/A'),
                'equipment_used': getattr(obj, 'equipment_used', 'N/A'),
                'property_removed': getattr(obj, 'property_removed', 'N/A'),
                'property_saved': getattr(obj, 'property_saved', 'N/A'),
                'property_lost': getattr(obj, 'property_lost', 'N/A'),
                'building_damage': getattr(obj, 'building_damage', 'N/A'),
                'items_destroyed': getattr(obj, 'items_destroyed', 'N/A'),
                'major_loss_cause': getattr(obj, 'major_loss_cause', 'N/A'),
                'assistance_details': '',
                'injured': ''
            }
        })
    
    # Process general incidents
    for obj in qs_general:
        all_reports.append({
            'type': 'General Incident',
            'date': obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            'sort_date': obj.form_date if obj.form_date else datetime.min.date(),
            'data': {
                'userName': obj.userName,
                'incident': getattr(obj, 'incident', 'N/A'),
                'call_number': getattr(obj, 'call_number', 'N/A'),
                'call_received': getattr(obj, 'call_received', 'N/A'),
                'time_left_station': getattr(obj, 'time_left_station', 'N/A'),
                'time_reached_scene': getattr(obj, 'time_reached_scene', 'N/A'),
                'time_returned': getattr(obj, 'time_returned', 'N/A'),
                'occupancy_type': getattr(obj, 'occupancy_type', 'N/A'),
                'construction_type': getattr(obj, 'construction_type', 'N/A'),
                'owner_details': getattr(obj, 'owner_details', 'N/A'),
                'electrical_chemicals': getattr(obj, 'electrical_chemicals', 'N/A'),
                'hazardous_materials': getattr(obj, 'hazardous_materials', 'N/A'),
                'weather_conditions': getattr(obj, 'weather_conditions', 'N/A'),
                'caller_contact': getattr(obj, 'caller_contact', 'N/A'),
                'source_of_call': getattr(obj, 'source_of_call', 'N/A'),
                'nature_incident': getattr(obj, 'nature_incident', 'N/A'),
                'exact_location': getattr(obj, 'exact_location', 'N/A'),
                'owner_occupant': getattr(obj, 'owner_occupant', 'N/A'),
                'premises_occupancy': getattr(obj, 'premises_occupancy', 'N/A'),
                'building_details': getattr(obj, 'building_details', 'N/A'),
                'premises_chemicals': getattr(obj, 'premises_chemicals', 'N/A'),
                'premises_hazardous': getattr(obj, 'premises_hazardous', 'N/A'),
                'deaths_male': getattr(obj, 'deaths_male', 'N/A'),
                'deaths_female': getattr(obj, 'deaths_female', 'N/A'),
                'deaths_adult': getattr(obj, 'deaths_adult', 'N/A'),
                'deaths_child': getattr(obj, 'deaths_child', 'N/A'),
                'injuries_male': '',
                'injuries_female': '',
                'injuries_adult': '',
                'injuries_child': '',
                'rescued_male': getattr(obj, 'rescued_male', 'N/A'),
                'rescued_female': getattr(obj, 'rescued_female', 'N/A'),
                'rescued_adult': getattr(obj, 'rescued_adult', 'N/A'),
                'rescued_child': getattr(obj, 'rescued_child', 'N/A'),
                'fire_personnel_injuries': getattr(obj, 'fire_personnel_injuries', 'N/A'),
                'animals_rescued': getattr(obj, 'animals_rescued', 'N/A'),
                'animals_lost': getattr(obj, 'animals_lost', 'N/A'),
                'hospital_doctor': getattr(obj, 'hospital_doctor', 'N/A'),
                'ambulance_notified': getattr(obj, 'ambulance_notified', 'N/A'),
                'appliances_crews': getattr(obj, 'appliances_crews', 'N/A'),
                'officer_in_charge': getattr(obj, 'officer_in_charge', 'N/A'),
                'equipment_used': getattr(obj, 'equipment_used', 'N/A'),
                'property_removed': getattr(obj, 'property_removed', 'N/A'),
                'property_saved': getattr(obj, 'property_saved', 'N/A'),
                'property_lost': getattr(obj, 'property_lost', 'N/A'),
                'building_damage': getattr(obj, 'building_damage', 'N/A'),
                'items_destroyed': getattr(obj, 'items_destroyed', 'N/A'),
                'major_loss_cause': getattr(obj, 'major_loss_cause', 'N/A'),
                'assistance_details': '',
                'injured': getattr(obj, 'injured', 'N/A')
            }
        })
    
    # Process assistance calls
    for obj in qs_assist:
        all_reports.append({
            'type': 'Assistance Call',
            'date': obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            'sort_date': obj.form_date if obj.form_date else datetime.min.date(),
            'data': {
                'userName': obj.userName,
                'incident': getattr(obj, 'incident_type', 'N/A'),
                'call_number': getattr(obj, 'call_number', 'N/A'),
                'call_received': getattr(obj, 'call_received', 'N/A'),
                'time_left_station': getattr(obj, 'time_left_station', 'N/A'),
                'time_reached_scene': getattr(obj, 'time_reached_scene', 'N/A'),
                'time_returned': getattr(obj, 'time_returned', 'N/A'),
                'occupancy_type': '',
                'construction_type': '',
                'owner_details': '',
                'electrical_chemicals': getattr(obj, 'electrical_chemicals', 'N/A'),
                'hazardous_materials': getattr(obj, 'hazardous_materials', 'N/A'),
                'weather_conditions': getattr(obj, 'weather_conditions', 'N/A'),
                'caller_contact': getattr(obj, 'caller_contact', 'N/A'),
                'source_of_call': getattr(obj, 'source_of_call', 'N/A'),
                'nature_incident': '',
                'exact_location': getattr(obj, 'exact_location', 'N/A'),
                'owner_occupant': '',
                'premises_occupancy': '',
                'building_details': '',
                'premises_chemicals': '',
                'premises_hazardous': '',
                'deaths_male': '',
                'deaths_female': '',
                'deaths_adult': '',
                'deaths_child': '',
                'injuries_male': '',
                'injuries_female': '',
                'injuries_adult': '',
                'injuries_child': '',
                'rescued_male': '',
                'rescued_female': '',
                'rescued_adult': '',
                'rescued_child': '',
                'fire_personnel_injuries': '',
                'animals_rescued': '',
                'animals_lost': '',
                'hospital_doctor': '',
                'ambulance_notified': '',
                'appliances_crews': '',
                'officer_in_charge': '',
                'equipment_used': '',
                'property_removed': '',
                'property_saved': '',
                'property_lost': '',
                'building_damage': '',
                'items_destroyed': '',
                'major_loss_cause': '',
                'assistance_details': getattr(obj, 'assistance_details', 'N/A'),
                'injured': ''
            }
        })
    
    # Sort by date (most recent first)
    all_reports = sorted(all_reports, key=lambda x: x['sort_date'], reverse=True)
    
    # Helper function to add labeled row
    def add_labeled_row(label, value, col_offset=0):
        nonlocal current_row
        label_col = get_column_letter(1 + col_offset)
        value_col = get_column_letter(2 + col_offset)
        
        label_cell = ws[f'{label_col}{current_row}']
        label_cell.value = label
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.alignment = left_alignment
        label_cell.border = thin_border
        
        value_cell = ws[f'{value_col}{current_row}']
        value_cell.value = value if value else 'N/A'
        value_cell.font = data_font
        value_cell.alignment = left_alignment
        value_cell.border = thin_border
        
        current_row += 1
    
    # Helper function to add section header
    def add_section_header(title):
        nonlocal current_row
        ws.merge_cells(f'A{current_row}:F{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = title
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center_alignment
        header_cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1
    
    # Write each report
    for report in all_reports:
        # Report separator with type and date
        ws.merge_cells(f'A{current_row}:F{current_row}')
        sep_cell = ws[f'A{current_row}']
        sep_cell.value = f"━━━ {report['type'].upper()} - {report['date']} ━━━"
        sep_cell.font = section_font
        sep_cell.fill = section_fill
        sep_cell.alignment = center_alignment
        sep_cell.border = thin_border
        ws.row_dimensions[current_row].height = 28
        current_row += 1
        
        # BASIC INFORMATION SECTION
        add_section_header('BASIC INFORMATION')
        add_labeled_row('Report Type', report['type'])
        add_labeled_row('Date of Filling the Form', report['date'])
        add_labeled_row('Name of the User', report['data']['userName'])
        add_labeled_row('Incident/Type', report['data']['incident'])
        add_labeled_row('Call Number', report['data']['call_number'])
        current_row += 1
        
        # TIMING DETAILS SECTION
        add_section_header('TIMING DETAILS')
        add_labeled_row('Call Received', report['data']['call_received'])
        add_labeled_row('Time Left Station', report['data']['time_left_station'])
        add_labeled_row('Time Reached Scene', report['data']['time_reached_scene'])
        add_labeled_row('Time Returned', report['data']['time_returned'])
        current_row += 1
        
        # LOCATION & PREMISES SECTION
        add_section_header('LOCATION & PREMISES INFORMATION')
        add_labeled_row('Exact Location', report['data']['exact_location'])
        add_labeled_row('Occupancy Type', report['data']['occupancy_type'])
        add_labeled_row('Construction Type', report['data']['construction_type'])
        add_labeled_row('Owner / Occupant', report['data']['owner_details'])
        add_labeled_row('Owner / Occupant Details', report['data']['owner_occupant'])
        add_labeled_row('Premises Occupancy Type', report['data']['premises_occupancy'])
        add_labeled_row('Building Details', report['data']['building_details'])
        current_row += 1
        
        # HAZARD INFORMATION SECTION
        add_section_header('HAZARD INFORMATION')
        add_labeled_row('Electrical / Gas / Chemicals', report['data']['electrical_chemicals'])
        add_labeled_row('Hazardous Materials', report['data']['hazardous_materials'])
        add_labeled_row('Premises Electrical / Gas / Chemicals', report['data']['premises_chemicals'])
        add_labeled_row('Premises Hazardous Materials', report['data']['premises_hazardous'])
        add_labeled_row('Weather Conditions', report['data']['weather_conditions'])
        current_row += 1
        
        # CALL INFORMATION SECTION
        add_section_header('CALL INFORMATION')
        add_labeled_row('Caller (Name & Contact)', report['data']['caller_contact'])
        add_labeled_row('Source of Call', report['data']['source_of_call'])
        add_labeled_row('Nature of Incident', report['data']['nature_incident'])
        current_row += 1
        
        # CASUALTIES SECTION
        add_section_header('CASUALTIES & INJURIES')
        add_labeled_row('Deaths Male', report['data']['deaths_male'])
        add_labeled_row('Deaths Female', report['data']['deaths_female'])
        add_labeled_row('Deaths Adult (Above 18)', report['data']['deaths_adult'])
        add_labeled_row('Deaths Child (Below 18)', report['data']['deaths_child'])
        add_labeled_row('Injuries Male', report['data']['injuries_male'])
        add_labeled_row('Injuries Female', report['data']['injuries_female'])
        add_labeled_row('Injuries Adult (Above 18)', report['data']['injuries_adult'])
        add_labeled_row('Injuries Child (Below 18)', report['data']['injuries_child'])
        add_labeled_row('Injured (GeneralIncident)', report['data']['injured'])
        current_row += 1
        
        # RESCUE INFORMATION SECTION
        add_section_header('RESCUE INFORMATION')
        add_labeled_row('Rescued Male', report['data']['rescued_male'])
        add_labeled_row('Rescued Female', report['data']['rescued_female'])
        add_labeled_row('Rescued Adult (Above 18)', report['data']['rescued_adult'])
        add_labeled_row('Rescued Child (Below 18)', report['data']['rescued_child'])
        add_labeled_row('Fire Personnel Injuries', report['data']['fire_personnel_injuries'])
        add_labeled_row('Animals Rescued', report['data']['animals_rescued'])
        add_labeled_row('Animals Lost', report['data']['animals_lost'])
        current_row += 1
        
        # EMERGENCY RESPONSE SECTION
        add_section_header('EMERGENCY RESPONSE')
        add_labeled_row('Hospital / Doctor', report['data']['hospital_doctor'])
        add_labeled_row('Ambulance / Police Notified Time', report['data']['ambulance_notified'])
        add_labeled_row('Appliances & Crews Attended', report['data']['appliances_crews'])
        add_labeled_row('Officer In Charge', report['data']['officer_in_charge'])
        add_labeled_row('Equipment Used', report['data']['equipment_used'])
        current_row += 1
        
        # PROPERTY & DAMAGE SECTION
        add_section_header('PROPERTY & DAMAGE ASSESSMENT')
        add_labeled_row('Property Removed', report['data']['property_removed'])
        add_labeled_row('Property Saved', report['data']['property_saved'])
        add_labeled_row('Property Lost', report['data']['property_lost'])
        add_labeled_row('Building Damage', report['data']['building_damage'])
        add_labeled_row('Items Destroyed', report['data']['items_destroyed'])
        add_labeled_row('Cause of Major Loss', report['data']['major_loss_cause'])
        current_row += 1
        
        # ADDITIONAL INFORMATION SECTION
        add_section_header('ADDITIONAL INFORMATION')
        add_labeled_row('Assistance Details', report['data']['assistance_details'])
        current_row += 2
    
    # Add summary footer
    ws.merge_cells(f'A{current_row}:F{current_row}')
    summary_cell = ws[f'A{current_row}']
    summary_cell.value = f'Total Reports: {len(all_reports)} | Firewater: {len(qs_fire)} | General Incident: {len(qs_general)} | Assistance Call: {len(qs_assist)}'
    summary_cell.font = Font(name='Arial', size=10, bold=True, italic=True)
    summary_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    summary_cell.alignment = center_alignment
    summary_cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reports_{datetime.now().strftime("%d%m_%H%M")}.xlsx"'
    
    wb.save(response)
    return response




def allcase_detail(request, model_type, case_id):
    if 'username' not in request.session:
        return redirect('admin_login')  # Should redirect to admin_login, not user_login
    
    username = request.session['username']
    
    # Fetch the case based on model type - WITHOUT filtering by userName
    if model_type == 'firewater':
        case = get_object_or_404(firewater, id=case_id)
    elif model_type == 'general':
        case = get_object_or_404(generalIncident, id=case_id)
    elif model_type == 'assistance':
        case = get_object_or_404(AssistanceCall, id=case_id)
    else:
        return redirect('adminviewreport')
    
    context = {
        'case': case,
        'model_type': model_type,
        'username': username,
        'is_admin': True,  # Flag to indicate this is admin view
    }
    return render(request, 'allcasedetail.html', context)





def adminanalytics(request):
    # Get date filters
    month = request.GET.get('month', '')
    start = request.GET.get('start_date', '')
    end = request.GET.get('end_date', '')

    # Prepare queryset filters for all 3 models
    q_fire = list(firewater.objects.all())
    q_general = list(generalIncident.objects.all())
    q_assist = list(AssistanceCall.objects.all())
    
    # Apply month filter
    if month:
        y, m = month.split('-')
        q_fire = [c for c in q_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        q_general = [c for c in q_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        q_assist = [c for c in q_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    # Apply start date filter
    if start:
        startdate = datetime.strptime(start, "%Y-%m-%d").date()
        q_fire = [c for c in q_fire if c.form_date >= startdate]
        q_general = [c for c in q_general if c.form_date >= startdate]
        q_assist = [c for c in q_assist if c.form_date >= startdate]
    
    # Apply end date filter
    if end:
        enddate = datetime.strptime(end, "%Y-%m-%d").date()
        q_fire = [c for c in q_fire if c.form_date <= enddate]
        q_general = [c for c in q_general if c.form_date <= enddate]
        q_assist = [c for c in q_assist if c.form_date <= enddate]
    
    # 1. Total cases - all reports in database
    total_cases = len(q_fire) + len(q_general) + len(q_assist)
    
    # 3. Assistance calls - total AssistanceCall reports
    assistance_calls = len(q_assist)
    
    # 4-11. Incident-specific counts from generalIncident
    vehicle_accidents = sum([c.incident.lower() == "vehicle accident" for c in q_general])
    well_related = sum([c.incident.lower() == "well related" for c in q_general])
    lift_accidents = sum([c.incident.lower() == "lift accidents" for c in q_general])
    quarry = sum([c.incident.lower() == "quarry" for c in q_general])
    landslides = sum([c.incident.lower() == "landslides" for c in q_general])
    building_collapse = sum([c.incident.lower() == "building collapse" for c in q_general])
    tree_related = sum([c.incident.lower() == "tree related" for c in q_general])
    animal_calls = sum([c.incident.lower() == "animal calls" for c in q_general])

    # 12. Animals rescued - sum from firewater and generalIncident
    animals_rescued = (sum([int(c.animals_rescued or 0) for c in q_fire]) +
                       sum([int(c.animals_rescued or 0) for c in q_general]))
    
    # 13. Animals lost - sum from firewater and generalIncident
    animals_lost = (sum([int(c.animals_lost or 0) for c in q_fire]) +
                    sum([int(c.animals_lost or 0) for c in q_general]))
    
    # property saved and lost
    fireproperty_saved = sum([
        int(c.property_saved or 0)
        for c in q_fire if c.incident == 'Fire'
    ])

    waterproperty_saved = sum([
        int(c.property_saved or 0)
        for c in q_fire if c.incident == 'Water'
    ])

    fireproperty_lost = sum([
        int(c.property_lost or 0)
        for c in q_fire if c.incident == 'Fire'
    ])

    waterproperty_lost = sum([
        int(c.property_lost or 0)
        for c in q_fire if c.incident == 'Water'
    ])


    # 14. Humans rescued - sum of rescued_male and rescued_female from both models
    human_rescued = (
        sum([int(c.rescued_male or 0) for c in q_fire]) +
        sum([int(c.rescued_female or 0) for c in q_fire]) +
        sum([int(c.rescued_male or 0) for c in q_general]) +
        sum([int(c.rescued_female or 0) for c in q_general])
    )
    
    # 15. Humans lost - sum of deaths_male and deaths_female from both models
    human_lost = (
        sum([int(c.deaths_male or 0) for c in q_fire]) +
        sum([int(c.deaths_female or 0) for c in q_fire]) +
        sum([int(c.deaths_male or 0) for c in q_general]) +
        sum([int(c.deaths_female or 0) for c in q_general])
    )
    
    # 16. Water related incidents - only from firewater where incident is "Water"
    # Filter water-related incidents (exact match with "Water")
    q_water = [c for c in q_fire if c.incident == 'Water']
    
    # Deaths - separate by gender and age
    water_deaths_male = sum([int(c.deaths_male or 0) for c in q_water])
    water_deaths_female = sum([int(c.deaths_female or 0) for c in q_water])
    water_deaths_adult = sum([int(c.deaths_adult or 0) for c in q_water])
    water_deaths_child = sum([int(c.deaths_child or 0) for c in q_water])
    
    # Rescued - separate by gender and age
    water_rescued_male = sum([int(c.rescued_male or 0) for c in q_water])
    water_rescued_female = sum([int(c.rescued_female or 0) for c in q_water])
    water_rescued_adult = sum([int(c.rescued_adult or 0) for c in q_water])
    water_rescued_child = sum([int(c.rescued_child or 0) for c in q_water])

    analysis = {
        'total_cases': total_cases,
        'assistance_calls': assistance_calls,
        'vehicle_accidents': vehicle_accidents,
        'well_related': well_related,
        'lift_accidents': lift_accidents,
        'quarry': quarry,
        'landslides': landslides,
        'building_collapse': building_collapse,
        'tree_related': tree_related,
        'animal_calls': animal_calls,
        'animals_rescued': animals_rescued,
        'animals_lost': animals_lost,
        'fireproperty_saved': fireproperty_saved,
        'waterproperty_saved': waterproperty_saved, 
        'fireproperty_lost': fireproperty_lost,
        'waterproperty_lost': waterproperty_lost,
        'human_rescued': human_rescued,
        'human_lost': human_lost,
        'water_deaths_male': water_deaths_male,
        'water_deaths_female': water_deaths_female,
        'water_deaths_adult': water_deaths_adult,
        'water_deaths_child': water_deaths_child,
        'water_rescued_male': water_rescued_male,
        'water_rescued_female': water_rescued_female,
        'water_rescued_adult': water_rescued_adult,
        'water_rescued_child': water_rescued_child,
    }

    return render(request, 'adminanalytics.html', {'analysis': analysis})


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
import io

def download_analytics_csv(request):
    """
    View to download analytics data as professionally formatted Excel file
    """
    # Get filter parameters
    month = request.GET.get('month', '')
    start = request.GET.get('start_date', '')
    end = request.GET.get('end_date', '')
    
    # Prepare queryset filters for all 3 models
    q_fire = list(firewater.objects.all())
    q_general = list(generalIncident.objects.all())
    q_assist = list(AssistanceCall.objects.all())
    
    # Apply month filter
    if month:
        y, m = month.split('-')
        q_fire = [c for c in q_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        q_general = [c for c in q_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        q_assist = [c for c in q_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    # Apply start date filter
    if start:
        startdate = datetime.strptime(start, "%Y-%m-%d").date()
        q_fire = [c for c in q_fire if c.form_date >= startdate]
        q_general = [c for c in q_general if c.form_date >= startdate]
        q_assist = [c for c in q_assist if c.form_date >= startdate]
    
    # Apply end date filter
    if end:
        enddate = datetime.strptime(end, "%Y-%m-%d").date()
        q_fire = [c for c in q_fire if c.form_date <= enddate]
        q_general = [c for c in q_general if c.form_date <= enddate]
        q_assist = [c for c in q_assist if c.form_date <= enddate]
    
    # Calculate statistics
    total_cases = len(q_fire) + len(q_general) + len(q_assist)
    assistance_calls = len(q_assist)
    
    vehicle_accidents = sum([c.incident.lower() == "vehicle accident" for c in q_general])
    well_related = sum([c.incident.lower() == "well related" for c in q_general])
    lift_accidents = sum([c.incident.lower() == "lift accidents" for c in q_general])
    quarry = sum([c.incident.lower() == "quarry" for c in q_general])
    landslides = sum([c.incident.lower() == "landslides" for c in q_general])
    building_collapse = sum([c.incident.lower() == "building collapse" for c in q_general])
    tree_related = sum([c.incident.lower() == "tree related" for c in q_general])
    animal_calls = sum([c.incident.lower() == "animal calls" for c in q_general])

    animals_rescued = (sum([int(c.animals_rescued or 0) for c in q_fire]) +
                       sum([int(c.animals_rescued or 0) for c in q_general]))
    animals_lost = (sum([int(c.animals_lost or 0) for c in q_fire]) +
                    sum([int(c.animals_lost or 0) for c in q_general]))

    human_rescued = (
        sum([int(c.rescued_male or 0) for c in q_fire]) +
        sum([int(c.rescued_female or 0) for c in q_fire]) +
        sum([int(c.rescued_male or 0) for c in q_general]) +
        sum([int(c.rescued_female or 0) for c in q_general])
    )
    human_lost = (
        sum([int(c.deaths_male or 0) for c in q_fire]) +
        sum([int(c.deaths_female or 0) for c in q_fire]) +
        sum([int(c.deaths_male or 0) for c in q_general]) +
        sum([int(c.deaths_female or 0) for c in q_general])
    )
    
    # Water related incidents
    q_water = [c for c in q_fire if 'water' in c.incident.lower()]
    qfire = [c for c in q_fire if c.incident and c.incident.lower() == 'fire']
    
    water_deaths_male = sum([int(c.deaths_male or 0) for c in q_water])
    water_deaths_female = sum([int(c.deaths_female or 0) for c in q_water])
    water_deaths_adult = sum([int(c.deaths_adult or 0) for c in q_water])
    water_deaths_child = sum([int(c.deaths_child or 0) for c in q_water])
    
    water_rescued_male = sum([int(c.rescued_male or 0) for c in q_water])
    water_rescued_female = sum([int(c.rescued_female or 0) for c in q_water])
    water_rescued_adult = sum([int(c.rescued_adult or 0) for c in q_water])
    water_rescued_child = sum([int(c.rescued_child or 0) for c in q_water])

    # Property saved and lost - CORRECTED
    fireproperty_saved = sum([
        int(c.property_saved or 0)
        for c in q_fire if c.incident and c.incident.lower() == 'fire'
    ])

    waterproperty_saved = sum([
        int(c.property_saved or 0)
        for c in q_fire if c.incident and 'water' in c.incident.lower()
    ])

    fireproperty_lost = sum([
        int(c.property_lost or 0)
        for c in q_fire if c.incident and c.incident.lower() == 'fire'
    ])

    waterproperty_lost = sum([
        int(c.property_lost or 0)
        for c in q_fire if c.incident and 'water' in c.incident.lower()
    ])
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Report"
    
    # Define styles
    title_font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    
    section_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
    
    subsection_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    subsection_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    
    row = 1
    
    # ========== MAIN TITLE ==========
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'INCIDENT ANALYTICS REPORT'
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 35
    row += 1
    
    # Report info section
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = 'Report Generated:'
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = left_align
    ws[f'C{row}'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    ws[f'C{row}'].font = data_font
    ws[f'C{row}'].alignment = left_align
    row += 1
    
    # Filter information
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = 'Report Period:'
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = left_align
    
    if month:
        month_name = datetime.strptime(month + '-01', '%Y-%m-%d').strftime('%B %Y')
        filter_text = month_name
    elif start and end:
        filter_text = f'{start} to {end}'
    elif start:
        filter_text = f'From {start} onwards'
    elif end:
        filter_text = f'Up to {end}'
    else:
        filter_text = 'All Records'
    
    ws[f'C{row}'] = filter_text
    ws[f'C{row}'].font = data_font
    ws[f'C{row}'].alignment = left_align
    row += 2
    
    # ========== SECTION 1: OVERVIEW ==========
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'SECTION 1: OVERALL STATISTICS'
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Table headers
    headers = ['Metric', 'Count']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    # Data rows
    overview_data = [
        ['Total Incidents', total_cases],
        ['Assistance Calls', assistance_calls],
        ['Fire Incidents', len(qfire)],
        ['Water Incidents', len(q_water)],
        ['General Incidents', len(q_general)]
    ]
    
    for idx, data_row in enumerate(overview_data):
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if idx % 2 == 1:
                cell.fill = alt_fill
        row += 1
    row += 1
    
    # ========== SECTION 2: INCIDENT CATEGORIES ==========
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'SECTION 2: INCIDENT BREAKDOWN BY CATEGORY'
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Table headers
    headers = ['Incident Type', 'Number of Cases']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    # Data rows
    incident_data = [
        ['Vehicle Accidents', vehicle_accidents],
        ['Well Related Incidents', well_related],
        ['Lift Accidents', lift_accidents],
        ['Quarry Incidents', quarry],
        ['Landslides', landslides],
        ['Building Collapse', building_collapse],
        ['Tree Related Incidents', tree_related],
        ['Animal Related Calls', animal_calls]
    ]
    
    for idx, data_row in enumerate(incident_data):
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if idx % 2 == 1:
                cell.fill = alt_fill
        row += 1
    row += 1


        # ========== SECTION 3: PROPERTIES CATEGORIES ==========
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'SECTION 3: PROPERTY ASSESSMENT'
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Table headers
    headers = ['Incident Type', 'Total Property Saved', 'Total Property Lost']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    # Data rows
    incident_data = [
        ['Fire', fireproperty_saved, fireproperty_lost],
        ['Water', waterproperty_saved, waterproperty_lost],
    ]
    
    for idx, data_row in enumerate(incident_data):
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if idx % 2 == 1:
                cell.fill = alt_fill
        row += 1
    row += 1
    
    # ========== SECTION 4: RESCUE & CASUALTY STATISTICS ==========
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'SECTION 4: RESCUE & CASUALTY STATISTICS'
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Table headers
    headers = ['Category', 'Rescued', 'Lost/Deceased', 'Survival Rate']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    # Data rows
    human_total = human_rescued + human_lost
    animal_total = animals_rescued + animals_lost
    
    rescue_data = [
        ['Humans', human_rescued, human_lost, f'{(human_rescued/human_total*100):.1f}%' if human_total > 0 else 'N/A'],
        ['Animals', animals_rescued, animals_lost, f'{(animals_rescued/animal_total*100):.1f}%' if animal_total > 0 else 'N/A']
    ]
    
    for idx, data_row in enumerate(rescue_data):
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if idx % 2 == 1:
                cell.fill = alt_fill
        row += 1
    row += 1
    
    # ========== SECTION 5: WATER RELATED INCIDENTS ==========
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'SECTION 5: WATER RELATED INCIDENTS'
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Summary info
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = 'Total Water Related Incidents:'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = left_align
    ws[f'C{row}'] = len(q_water)
    ws[f'C{row}'].font = Font(name='Arial', size=11, bold=True)
    ws[f'C{row}'].alignment = center_align
    row += 2
    
    # Subsection: Deaths
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = '4.1 CASUALTIES (Deaths)'
    cell.font = subsection_font
    cell.fill = subsection_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 22
    row += 1
    
    # Table headers
    headers = ['Category', 'Count', 'Percentage', 'Age Group']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    total_water_deaths = water_deaths_male + water_deaths_female
    
    death_data = [
        ['Male Deaths', water_deaths_male, f'{(water_deaths_male/total_water_deaths*100):.1f}%' if total_water_deaths > 0 else '0%', 'All Ages'],
        ['Female Deaths', water_deaths_female, f'{(water_deaths_female/total_water_deaths*100):.1f}%' if total_water_deaths > 0 else '0%', 'All Ages'],
        ['Adult Deaths', water_deaths_adult, f'{(water_deaths_adult/total_water_deaths*100):.1f}%' if total_water_deaths > 0 else '0%', '18+ years'],
        ['Child Deaths', water_deaths_child, f'{(water_deaths_child/total_water_deaths*100):.1f}%' if total_water_deaths > 0 else '0%', '<18 years'],
        ['TOTAL DEATHS', total_water_deaths]
    ]
    
    for idx, data_row in enumerate(death_data):
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if idx == len(death_data) - 1:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if idx % 2 == 1 and idx != len(death_data) - 1:
                cell.fill = alt_fill
        row += 1
    row += 1
    
    # Subsection: Rescues
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = '4.2 SUCCESSFUL RESCUES'
    cell.font = subsection_font
    cell.fill = subsection_fill
    cell.alignment = center_align
    ws.row_dimensions[row].height = 22
    row += 1
    
    # Table headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    
    total_water_rescued = water_rescued_male + water_rescued_female
    
    rescue_data = [
        ['Male Rescued', water_rescued_male, f'{(water_rescued_male/total_water_rescued*100):.1f}%' if total_water_rescued > 0 else '0%', 'All Ages'],
        ['Female Rescued', water_rescued_female, f'{(water_rescued_female/total_water_rescued*100):.1f}%' if total_water_rescued > 0 else '0%', 'All Ages'],
        ['Adults Rescued', water_rescued_adult, f'{(water_rescued_adult/total_water_rescued*100):.1f}%' if total_water_rescued > 0 else '0%', '18+ years'],
        ['Children Rescued', water_rescued_child, f'{(water_rescued_child/total_water_rescued*100):.1f}%' if total_water_rescued > 0 else '0%', '<18 years'],
        ['TOTAL RESCUED', total_water_rescued]
    ]
    
    for idx, data_row in enumerate(rescue_data):
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if idx == len(rescue_data) - 1:
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.font = data_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if idx % 2 == 1 and idx != len(rescue_data) - 1:
                cell.fill = alt_fill
        row += 1
    row += 2
    
    # Footer
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━'
    cell.alignment = center_align
    row += 1
    
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'END OF REPORT'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = center_align
    row += 1
    
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = 'This report contains statistical data for the specified time period. For detailed incident records, please refer to individual case files.'
    cell.font = Font(name='Arial', size=9, italic=True)
    cell.alignment = center_align
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.now().strftime('%d%m_%H%M')
    filename = f'Incident_Analytics_Report_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response




def adminadduser(request):
    # ✅ FIXED: Session check
    if 'username' not in request.session:
        return redirect('admin_login')  # Correct admin redirect
    
    if request.method == "POST":
        form_type = request.POST.get("form_type")
        
        if form_type == "user":
            fullname = request.POST.get("fullname")
            username = request.POST.get("username")
            userid = request.POST.get("userid")
            password = request.POST.get("password")
            confirmpassword = request.POST.get("confirmpassword")
            
            # ✅ ADDED: Input validation
            if not all([fullname, username, userid, password, confirmpassword]):
                messages.error(request, "All fields are required.")
                return redirect("adminadduser")
            
            # ✅ ADDED: Check for duplicate username
            if userdetails.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.")
                return redirect("adminadduser")
            
            # ✅ ADDED: Check for duplicate user ID
            if userdetails.objects.filter(userid=userid).exists():
                messages.error(request, "User ID already exists.")
                return redirect("adminadduser")
            
            if password != confirmpassword:
                messages.error(request, "User passwords do not match.")
                return redirect("adminadduser")
            
            # ✅ SECURITY: Hash passwords before storing
            # You should use Django's built-in password hashing
            # from django.contrib.auth.hashers import make_password
            # hashed_password = make_password(password)
            
            try:
                userdetails.objects.create(
                    fullname=fullname,
                    username=username,
                    userid=userid,
                    password=password,  # Should be hashed!
                    confirmpassword=confirmpassword
                )
                messages.success(request, f"User {username} added successfully.")
            except Exception as e:
                messages.error(request, f"Error adding user: {str(e)}")
            
            return redirect("adminadduser")
        
        elif form_type == "admin":
            adminname = request.POST.get("adminname")
            adminusername = request.POST.get("adminusername")
            adminid = request.POST.get("adminid")
            adminpassword = request.POST.get("adminpassword")
            adminconfirmpassword = request.POST.get("adminconfirmpassword")
            
            # ✅ ADDED: Input validation
            if not all([adminname, adminusername, adminid, adminpassword, adminconfirmpassword]):
                messages.error(request, "All fields are required.")
                return redirect("adminadduser")
            
            # ✅ ADDED: Check for duplicate username
            if admindetails.objects.filter(adminusername=adminusername).exists():
                messages.error(request, "Admin username already exists.")
                return redirect("adminadduser")
            
            # ✅ ADDED: Check for duplicate admin ID
            if admindetails.objects.filter(adminid=adminid).exists():
                messages.error(request, "Admin ID already exists.")
                return redirect("adminadduser")
            
            if adminpassword != adminconfirmpassword:
                messages.error(request, "Admin passwords do not match.")
                return redirect("adminadduser")
            
            try:
                admindetails.objects.create(
                    adminname=adminname,
                    adminusername=adminusername,
                    adminid=adminid,
                    adminpassword=adminpassword,  # Should be hashed!
                    adminconfirmpassword=adminconfirmpassword
                )
                messages.success(request, f"Admin {adminusername} added successfully.")
            except Exception as e:
                messages.error(request, f"Error adding admin: {str(e)}")
            
            return redirect("adminadduser")
    
    # GET request
    all_users = userdetails.objects.all()
    all_admins = admindetails.objects.all()
    
    return render(request, "adminaddusers.html", {
        "all_users": all_users,
        "all_admins": all_admins,
    })

def deleteuser(request, user_id):
    # ✅ ADDED: Session check
    if 'username' not in request.session:
        return redirect('admin_login')
    
    # ✅ ADDED: POST method check for security
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("adminadduser")
    
    try:
        user = userdetails.objects.get(id=user_id)
        username = user.username
        user.delete()
        messages.success(request, f"User {username} deleted successfully.")
    except userdetails.DoesNotExist:
        messages.error(request, "User not found.")
    except Exception as e:
        messages.error(request, f"Error deleting user: {str(e)}")
    
    return redirect("adminadduser")

def deleteadmin(request, admin_id):
    # ✅ ADDED: Session check
    if 'username' not in request.session:
        return redirect('admin_login')
    
    # ✅ ADDED: POST method check for security
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("adminadduser")
    
    # ✅ ADDED: Prevent deleting yourself
    try:
        admin = admindetails.objects.get(id=admin_id)
        
        # Check if trying to delete current logged-in admin
        if admin.adminusername == request.session.get('username'):
            messages.error(request, "You cannot delete your own admin account.")
            return redirect("adminadduser")
        
        adminusername = admin.adminusername
        admin.delete()
        messages.success(request, f"Admin {adminusername} deleted successfully.")
    except admindetails.DoesNotExist:
        messages.error(request, "Admin not found.")
    except Exception as e:
        messages.error(request, f"Error deleting admin: {str(e)}")
    
    return redirect("adminadduser")