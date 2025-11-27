from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import redirect

def download_reports(request):
    if 'username' not in request.session:
        return redirect('admin_login')
    
    # Get filtered reports
    qs_fire = list(firewater.objects.all())
    qs_general = list(generalIncident.objects.all())
    qs_assist = list(AssistanceCall.objects.all())

    month = request.GET.get('month')
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    
    # Apply filters
    if month:
        y, m = month.split('-')
        qs_fire = [c for c in qs_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_general = [c for c in qs_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_assist = [c for c in qs_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    if start:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date >= start_date]
        qs_general = [c for c in qs_general if c.form_date >= start_date]
        qs_assist = [c for c in qs_assist if c.form_date >= start_date]
    
    if end:
        end_date = datetime.strptime(end, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date <= end_date]
        qs_general = [c for c in qs_general if c.form_date <= end_date]
        qs_assist = [c for c in qs_assist if c.form_date <= end_date]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Incident Reports"
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    
    section_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Row counter
    current_row = 1
    
    # Main Title
    ws.merge_cells(f'A{current_row}:E{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = 'FIRE & RESCUE SERVICES - INCIDENT REPORTS'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Report Details
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].value = f'Report Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}'
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, italic=True)
    ws[f'A{current_row}'].alignment = left_alignment
    
    if month:
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'].value = f'Period: {month}'
        ws[f'C{current_row}'].font = Font(name='Arial', size=10, italic=True)
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    elif start and end:
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'].value = f'Period: {start} to {end}'
        ws[f'C{current_row}'].font = Font(name='Arial', size=10, italic=True)
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    
    current_row += 2
    
    # Section structure with headers
    sections = [
        {
            'name': 'BASIC INFORMATION',
            'headers': ['Report Type', 'Date', 'User Name', 'Incident/Type', 'Call Number']
        },
        {
            'name': 'TIMING DETAILS',
            'headers': ['Call Received', 'Time Left Station', 'Time Reached Scene', 'Time Returned']
        },
        {
            'name': 'LOCATION & PREMISES',
            'headers': ['Exact Location', 'Occupancy Type', 'Construction Type', 'Building Details', 'Owner/Occupant']
        },
        {
            'name': 'HAZARD INFORMATION',
            'headers': ['Electrical/Gas/Chemicals', 'Hazardous Materials', 'Weather Conditions']
        },
        {
            'name': 'CALL INFORMATION',
            'headers': ['Caller (Name & Contact)', 'Source of Call', 'Nature of Incident']
        },
        {
            'name': 'CASUALTIES & RESCUE',
            'headers': ['Deaths (M/F/Adult/Child)', 'Injuries (M/F/Adult/Child)', 'Rescued (M/F/Adult/Child)', 
                       'Fire Personnel Injuries', 'Animals Rescued', 'Animals Lost']
        },
        {
            'name': 'EMERGENCY RESPONSE',
            'headers': ['Hospital/Doctor', 'Ambulance/Police Notified', 'Appliances & Crews', 
                       'Officer In Charge', 'Equipment Used']
        },
        {
            'name': 'PROPERTY & DAMAGE',
            'headers': ['Property Removed', 'Property Saved', 'Property Lost', 'Building Damage', 
                       'Items Destroyed', 'Cause of Major Loss']
        },
        {
            'name': 'ADDITIONAL INFORMATION',
            'headers': ['Assistance Details', 'Other Notes']
        }
    ]
    
    # Prepare all data
    all_reports = []
    
    # Process firewater reports
    for obj in qs_fire:
        all_reports.append({
            'type': 'Firewater',
            'date': obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            'data': {
                'userName': obj.userName,
                'incident': getattr(obj, 'incident', 'N/A'),
                'call_number': getattr(obj, 'call_number', 'N/A'),
                'call_received': getattr(obj, 'call_received', 'N/A'),
                'time_left_station': getattr(obj, 'time_left_station', 'N/A'),
                'time_reached_scene': getattr(obj, 'time_reached_scene', 'N/A'),
                'time_returned': getattr(obj, 'time_returned', 'N/A'),
                'exact_location': getattr(obj, 'exact_location', 'N/A'),
                'occupancy_type': getattr(obj, 'occupancy_type', 'N/A'),
                'construction_type': getattr(obj, 'construction_type', 'N/A'),
                'building_details': getattr(obj, 'building_details', 'N/A'),
                'owner_details': getattr(obj, 'owner_details', 'N/A'),
                'electrical_chemicals': getattr(obj, 'electrical_chemicals', 'N/A'),
                'hazardous_materials': getattr(obj, 'hazardous_materials', 'N/A'),
                'weather_conditions': getattr(obj, 'weather_conditions', 'N/A'),
                'caller_contact': getattr(obj, 'caller_contact', 'N/A'),
                'source_of_call': getattr(obj, 'source_of_call', 'N/A'),
                'nature_incident': getattr(obj, 'nature_incident', 'N/A'),
                'casualties': f"D: {getattr(obj, 'deaths_male', 0)}/{getattr(obj, 'deaths_female', 0)}/{getattr(obj, 'deaths_adult', 0)}/{getattr(obj, 'deaths_child', 0)}",
                'injuries': f"I: {getattr(obj, 'injuries_male', 0)}/{getattr(obj, 'injuries_female', 0)}/{getattr(obj, 'injuries_adult', 0)}/{getattr(obj, 'injuries_child', 0)}",
                'rescued': f"R: {getattr(obj, 'rescued_male', 0)}/{getattr(obj, 'rescued_female', 0)}/{getattr(obj, 'rescued_adult', 0)}/{getattr(obj, 'rescued_child', 0)}",
                'fire_personnel_injuries': getattr(obj, 'fire_personnel_injuries', 'N/A'),
                'animals_rescued': getattr(obj, 'animals_rescued', 'N/A'),
                'animals_lost': getattr(obj, 'animals_lost', 'N/A'),
                'hospital_doctor': getattr(obj, 'hospital_doctor', 'N/A'),
                'ambulance_notified': getattr(obj, 'ambulance_notified', 'N/A'),
                'appliances_crews': getattr(obj, 'appliances_crews', 'N/A'),
                'officer_in_charge': getattr(obj, 'officer_in_charge', 'N/A'),
                'equipment_used': getattr(obj, 'equipment_used', 'N/A'),
                'property_removed': getattr(obj, 'property_removed', 'N/A'),
                'property_saved': getattr(obj, 'property_saved', 'N/A'),
                'property_lost': getattr(obj, 'property_lost', 'N/A'),
                'building_damage': getattr(obj, 'building_damage', 'N/A'),
                'items_destroyed': getattr(obj, 'items_destroyed', 'N/A'),
                'major_loss_cause': getattr(obj, 'major_loss_cause', 'N/A'),
                'assistance_details': '',
                'other_notes': ''
            }
        })
    
    # Process general incidents
    for obj in qs_general:
        all_reports.append({
            'type': 'General Incident',
            'date': obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            'data': {
                'userName': obj.userName,
                'incident': getattr(obj, 'incident', 'N/A'),
                'call_number': getattr(obj, 'call_number', 'N/A'),
                'call_received': getattr(obj, 'call_received', 'N/A'),
                'time_left_station': getattr(obj, 'time_left_station', 'N/A'),
                'time_reached_scene': getattr(obj, 'time_reached_scene', 'N/A'),
                'time_returned': getattr(obj, 'time_returned', 'N/A'),
                'exact_location': getattr(obj, 'exact_location', 'N/A'),
                'occupancy_type': getattr(obj, 'occupancy_type', 'N/A'),
                'construction_type': getattr(obj, 'construction_type', 'N/A'),
                'building_details': getattr(obj, 'building_details', 'N/A'),
                'owner_details': getattr(obj, 'owner_details', 'N/A'),
                'electrical_chemicals': getattr(obj, 'electrical_chemicals', 'N/A'),
                'hazardous_materials': getattr(obj, 'hazardous_materials', 'N/A'),
                'weather_conditions': getattr(obj, 'weather_conditions', 'N/A'),
                'caller_contact': getattr(obj, 'caller_contact', 'N/A'),
                'source_of_call': getattr(obj, 'source_of_call', 'N/A'),
                'nature_incident': getattr(obj, 'nature_incident', 'N/A'),
                'casualties': f"D: {getattr(obj, 'deaths_male', 0)}/{getattr(obj, 'deaths_female', 0)}/{getattr(obj, 'deaths_adult', 0)}/{getattr(obj, 'deaths_child', 0)}",
                'injuries': '',
                'rescued': f"R: {getattr(obj, 'rescued_male', 0)}/{getattr(obj, 'rescued_female', 0)}/{getattr(obj, 'rescued_adult', 0)}/{getattr(obj, 'rescued_child', 0)}",
                'fire_personnel_injuries': getattr(obj, 'fire_personnel_injuries', 'N/A'),
                'animals_rescued': getattr(obj, 'animals_rescued', 'N/A'),
                'animals_lost': getattr(obj, 'animals_lost', 'N/A'),
                'hospital_doctor': getattr(obj, 'hospital_doctor', 'N/A'),
                'ambulance_notified': getattr(obj, 'ambulance_notified', 'N/A'),
                'appliances_crews': getattr(obj, 'appliances_crews', 'N/A'),
                'officer_in_charge': getattr(obj, 'officer_in_charge', 'N/A'),
                'equipment_used': getattr(obj, 'equipment_used', 'N/A'),
                'property_removed': getattr(obj, 'property_removed', 'N/A'),
                'property_saved': getattr(obj, 'property_saved', 'N/A'),
                'property_lost': getattr(obj, 'property_lost', 'N/A'),
                'building_damage': getattr(obj, 'building_damage', 'N/A'),
                'items_destroyed': getattr(obj, 'items_destroyed', 'N/A'),
                'major_loss_cause': getattr(obj, 'major_loss_cause', 'N/A'),
                'assistance_details': '',
                'other_notes': getattr(obj, 'injured', 'N/A')
            }
        })
    
    # Process assistance calls
    for obj in qs_assist:
        all_reports.append({
            'type': 'Assistance Call',
            'date': obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            'data': {
                'userName': obj.userName,
                'incident': getattr(obj, 'incident_type', 'N/A'),
                'call_number': getattr(obj, 'call_number', 'N/A'),
                'call_received': getattr(obj, 'call_received', 'N/A'),
                'time_left_station': getattr(obj, 'time_left_station', 'N/A'),
                'time_reached_scene': getattr(obj, 'time_reached_scene', 'N/A'),
                'time_returned': getattr(obj, 'time_returned', 'N/A'),
                'exact_location': getattr(obj, 'exact_location', 'N/A'),
                'occupancy_type': '',
                'construction_type': '',
                'building_details': '',
                'owner_details': '',
                'electrical_chemicals': getattr(obj, 'electrical_chemicals', 'N/A'),
                'hazardous_materials': getattr(obj, 'hazardous_materials', 'N/A'),
                'weather_conditions': getattr(obj, 'weather_conditions', 'N/A'),
                'caller_contact': getattr(obj, 'caller_contact', 'N/A'),
                'source_of_call': getattr(obj, 'source_of_call', 'N/A'),
                'nature_incident': '',
                'casualties': '',
                'injuries': '',
                'rescued': '',
                'fire_personnel_injuries': '',
                'animals_rescued': '',
                'animals_lost': '',
                'hospital_doctor': '',
                'ambulance_notified': '',
                'appliances_crews': '',
                'officer_in_charge': '',
                'equipment_used': '',
                'property_removed': '',
                'property_saved': '',
                'property_lost': '',
                'building_damage': '',
                'items_destroyed': '',
                'major_loss_cause': '',
                'assistance_details': getattr(obj, 'assistance_details', 'N/A'),
                'other_notes': ''
            }
        })
    
    # Sort by date
    all_reports = sorted(all_reports, key=lambda x: x['date'], reverse=True)
    
    # Write each report
    for report in all_reports:
        # Report separator
        ws.merge_cells(f'A{current_row}:E{current_row}')
        sep_cell = ws[f'A{current_row}']
        sep_cell.value = f"{report['type']} - {report['date']}"
        sep_cell.font = section_font
        sep_cell.fill = section_fill
        sep_cell.alignment = center_alignment
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Basic Information
        ws[f'A{current_row}'] = 'Report Type'
        ws[f'B{current_row}'] = report['type']
        ws[f'C{current_row}'] = 'Date'
        ws[f'D{current_row}'] = report['date']
        current_row += 1
        
        ws[f'A{current_row}'] = 'User Name'
        ws[f'B{current_row}'] = report['data']['userName']
        ws[f'C{current_row}'] = 'Incident Type'
        ws[f'D{current_row}'] = report['data']['incident']
        current_row += 1
        
        ws[f'A{current_row}'] = 'Call Number'
        ws[f'B{current_row}'] = report['data']['call_number']
        current_row += 2
        
        # Timing Details section
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'].value = 'TIMING DETAILS'
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws[f'A{current_row}'] = 'Call Received'
        ws[f'B{current_row}'] = report['data']['call_received']
        ws[f'C{current_row}'] = 'Left Station'
        ws[f'D{current_row}'] = report['data']['time_left_station']
        current_row += 1
        
        ws[f'A{current_row}'] = 'Reached Scene'
        ws[f'B{current_row}'] = report['data']['time_reached_scene']
        ws[f'C{current_row}'] = 'Returned'
        ws[f'D{current_row}'] = report['data']['time_returned']
        current_row += 2
        
        # Add more sections as needed...
        # (Continue with other sections following the same pattern)
        
        current_row += 1
    
    # Apply borders and adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Incident_Reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response














def download_reports(request):
    if 'username' not in request.session:
        return redirect('admin_login')
    
    # Get filtered reports
    qs_fire = list(firewater.objects.all())
    qs_general = list(generalIncident.objects.all())
    qs_assist = list(AssistanceCall.objects.all())

    month = request.GET.get('month')
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    
    if month:
        y, m = month.split('-')
        qs_fire = [c for c in qs_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_general = [c for c in qs_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        qs_assist = [c for c in qs_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    if start:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date >= start_date]
        qs_general = [c for c in qs_general if c.form_date >= start_date]
        qs_assist = [c for c in qs_assist if c.form_date >= start_date]
    
    if end:
        end_date = datetime.strptime(end, '%Y-%m-%d').date()
        qs_fire = [c for c in qs_fire if c.form_date <= end_date]
        qs_general = [c for c in qs_general if c.form_date <= end_date]
        qs_assist = [c for c in qs_assist if c.form_date <= end_date]

    # Define comprehensive header
    header = [
        "Report Type",
        "Date of Filling the Form",
        "Name of the User",
        "Incident/Type",
        "Call Number",
        "Call Received",
        "Time Left Station",
        "Time Reached Scene",
        "Time Returned",
        "Occupancy Type",
        "Construction Type",
        "Owner / Occupant",
        "Electrical / Gas / Chemicals",
        "Hazardous Materials",
        "Weather Conditions",
        "Caller (Name & Contact)",
        "Source of Call",
        "Nature of Incident",
        "Exact Location",
        "Owner / Occupant Details",
        "Premises Occupancy Type",
        "Building Details",
        "Premises Electrical / Gas / Chemicals",
        "Premises Hazardous Materials",
        "Deaths Male",
        "Deaths Female",
        "Deaths Adult (Above 18)",
        "Deaths Child (Below 18)",
        "Injuries Male",
        "Injuries Female",
        "Injuries Adult (Above 18)",
        "Injuries Child (Below 18)",
        "Rescued Male",
        "Rescued Female",
        "Rescued Adult (Above 18)",
        "Rescued Child (Below 18)",
        "Fire Personnel Injuries",
        "Animals Rescued",
        "Animals Lost",
        "Hospital / Doctor",
        "Ambulance / Police Notified Time",
        "Appliances & Crews Attended",
        "Officer In Charge",
        "Equipment Used",
        "Property Removed",
        "Property Saved",
        "Property Lost",
        "Building Damage",
        "Items Destroyed",
        "Cause of Major Loss",
        "Assistance Details",
        "Injured (GeneralIncident)",
    ]

    # Build all reports with complete data
    all_reports = []
    
    # Add firewater reports
    for obj in qs_fire:
        all_reports.append([
            "Firewater",
            obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            obj.userName,
            getattr(obj, 'incident', 'N/A'),
            getattr(obj, 'call_number', 'N/A'),
            getattr(obj, 'call_received', 'N/A'),
            getattr(obj, 'time_left_station', 'N/A'),
            getattr(obj, 'time_reached_scene', 'N/A'),
            getattr(obj, 'time_returned', 'N/A'),
            getattr(obj, 'occupancy_type', 'N/A'),
            getattr(obj, 'construction_type', 'N/A'),
            getattr(obj, 'owner_details', 'N/A'),
            getattr(obj, 'electrical_chemicals', 'N/A'),
            getattr(obj, 'hazardous_materials', 'N/A'),
            getattr(obj, 'weather_conditions', 'N/A'),
            getattr(obj, 'caller_contact', 'N/A'),
            getattr(obj, 'source_of_call', 'N/A'),
            getattr(obj, 'nature_incident', 'N/A'),
            getattr(obj, 'exact_location', 'N/A'),
            getattr(obj, 'owner_occupant', 'N/A'),
            getattr(obj, 'premises_occupancy', 'N/A'),
            getattr(obj, 'building_details', 'N/A'),
            getattr(obj, 'premises_chemicals', 'N/A'),
            getattr(obj, 'premises_hazardous', 'N/A'),
            getattr(obj, 'deaths_male', 'N/A'),
            getattr(obj, 'deaths_female', 'N/A'),
            getattr(obj, 'deaths_adult', 'N/A'),
            getattr(obj, 'deaths_child', 'N/A'),
            getattr(obj, 'injuries_male', 'N/A'),
            getattr(obj, 'injuries_female', 'N/A'),
            getattr(obj, 'injuries_adult', 'N/A'),
            getattr(obj, 'injuries_child', 'N/A'),
            getattr(obj, 'rescued_male', 'N/A'),
            getattr(obj, 'rescued_female', 'N/A'),
            getattr(obj, 'rescued_adult', 'N/A'),
            getattr(obj, 'rescued_child', 'N/A'),
            getattr(obj, 'fire_personnel_injuries', 'N/A'),
            getattr(obj, 'animals_rescued', 'N/A'),
            getattr(obj, 'animals_lost', 'N/A'),
            getattr(obj, 'hospital_doctor', 'N/A'),
            getattr(obj, 'ambulance_notified', 'N/A'),
            getattr(obj, 'appliances_crews', 'N/A'),
            getattr(obj, 'officer_in_charge', 'N/A'),
            getattr(obj, 'equipment_used', 'N/A'),
            getattr(obj, 'property_removed', 'N/A'),
            getattr(obj, 'property_saved', 'N/A'),
            getattr(obj, 'property_lost', 'N/A'),
            getattr(obj, 'building_damage', 'N/A'),
            getattr(obj, 'items_destroyed', 'N/A'),
            getattr(obj, 'major_loss_cause', 'N/A'),
            '',  # Assistance Details (not applicable)
            '',  # Injured GeneralIncident (not applicable)
        ])
    
    # Add general incident reports
    for obj in qs_general:
        all_reports.append([
            "GeneralIncident",
            obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            obj.userName,
            getattr(obj, 'incident', 'N/A'),
            getattr(obj, 'call_number', 'N/A'),
            getattr(obj, 'call_received', 'N/A'),
            getattr(obj, 'time_left_station', 'N/A'),
            getattr(obj, 'time_reached_scene', 'N/A'),
            getattr(obj, 'time_returned', 'N/A'),
            getattr(obj, 'occupancy_type', 'N/A'),
            getattr(obj, 'construction_type', 'N/A'),
            getattr(obj, 'owner_details', 'N/A'),
            getattr(obj, 'electrical_chemicals', 'N/A'),
            getattr(obj, 'hazardous_materials', 'N/A'),
            getattr(obj, 'weather_conditions', 'N/A'),
            getattr(obj, 'caller_contact', 'N/A'),
            getattr(obj, 'source_of_call', 'N/A'),
            getattr(obj, 'nature_incident', 'N/A'),
            getattr(obj, 'exact_location', 'N/A'),
            getattr(obj, 'owner_occupant', 'N/A'),
            getattr(obj, 'premises_occupancy', 'N/A'),
            getattr(obj, 'building_details', 'N/A'),
            getattr(obj, 'premises_chemicals', 'N/A'),
            getattr(obj, 'premises_hazardous', 'N/A'),
            getattr(obj, 'deaths_male', 'N/A'),
            getattr(obj, 'deaths_female', 'N/A'),
            getattr(obj, 'deaths_adult', 'N/A'),
            getattr(obj, 'deaths_child', 'N/A'),
            '',  # Injuries male (not in GeneralIncident)
            '',  # Injuries female
            '',  # Injuries adult
            '',  # Injuries child
            getattr(obj, 'rescued_male', 'N/A'),
            getattr(obj, 'rescued_female', 'N/A'),
            getattr(obj, 'rescued_adult', 'N/A'),
            getattr(obj, 'rescued_child', 'N/A'),
            getattr(obj, 'fire_personnel_injuries', 'N/A'),
            getattr(obj, 'animals_rescued', 'N/A'),
            getattr(obj, 'animals_lost', 'N/A'),
            getattr(obj, 'hospital_doctor', 'N/A'),
            getattr(obj, 'ambulance_notified', 'N/A'),
            getattr(obj, 'appliances_crews', 'N/A'),
            getattr(obj, 'officer_in_charge', 'N/A'),
            getattr(obj, 'equipment_used', 'N/A'),
            getattr(obj, 'property_removed', 'N/A'),
            getattr(obj, 'property_saved', 'N/A'),
            getattr(obj, 'property_lost', 'N/A'),
            getattr(obj, 'building_damage', 'N/A'),
            getattr(obj, 'items_destroyed', 'N/A'),
            getattr(obj, 'major_loss_cause', 'N/A'),
            '',  # Assistance Details (not applicable)
            getattr(obj, 'injured', 'N/A'),  # GeneralIncident specific field
        ])
    
    # Add assistance call reports
    for obj in qs_assist:
        all_reports.append([
            "AssistanceCall",
            obj.form_date.strftime('%d-%m-%Y') if obj.form_date else 'N/A',
            obj.userName,
            getattr(obj, 'incident_type', 'N/A'),
            getattr(obj, 'call_number', 'N/A'),
            getattr(obj, 'call_received', 'N/A'),
            getattr(obj, 'time_left_station', 'N/A'),
            getattr(obj, 'time_reached_scene', 'N/A'),
            getattr(obj, 'time_returned', 'N/A'),
            '',  # Occupancy Type (not in AssistanceCall)
            '',  # Construction Type
            '',  # Owner / Occupant
            getattr(obj, 'electrical_chemicals', 'N/A'),
            getattr(obj, 'hazardous_materials', 'N/A'),
            getattr(obj, 'weather_conditions', 'N/A'),
            getattr(obj, 'caller_contact', 'N/A'),
            getattr(obj, 'source_of_call', 'N/A'),
            '',  # Nature of Incident
            getattr(obj, 'exact_location', 'N/A'),
            '',  # Owner / Occupant Details
            '',  # Premises Occupancy Type
            '',  # Building Details
            '',  # Premises Electrical / Gas / Chemicals
            '',  # Premises Hazardous Materials
            '',  # Deaths Male
            '',  # Deaths Female
            '',  # Deaths Adult
            '',  # Deaths Child
            '',  # Injuries Male
            '',  # Injuries Female
            '',  # Injuries Adult
            '',  # Injuries Child
            '',  # Rescued Male
            '',  # Rescued Female
            '',  # Rescued Adult
            '',  # Rescued Child
            '',  # Fire Personnel Injuries
            '',  # Animals Rescued
            '',  # Animals Lost
            '',  # Hospital / Doctor
            '',  # Ambulance / Police Notified Time
            '',  # Appliances & Crews Attended
            '',  # Officer In Charge
            '',  # Equipment Used
            '',  # Property Removed
            '',  # Property Saved
            '',  # Property Lost
            '',  # Building Damage
            '',  # Items Destroyed
            '',  # Cause of Major Loss
            getattr(obj, 'assistance_details', 'N/A'),
            '',  # Injured (GeneralIncident only)
        ])
    
    # Sort by date (most recent first)
    all_reports = sorted(all_reports, key=lambda c: c[1], reverse=True)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Monthly_reports.csv"'
    
    writer = csv.writer(response)
    writer.writerow(header)
    writer.writerows(all_reports)
    
    return response









def download_analytics_csv(request):
    """
    View to download analytics data as CSV with corrected logic
    """
    # Get filter parameters
    month = request.GET.get('month', '')
    start = request.GET.get('start_date', '')
    end = request.GET.get('end_date', '')
    
    # Prepare queryset filters for all 3 models
    q_fire = list(firewater.objects.all())
    q_general = list(generalIncident.objects.all())
    q_assist = list(AssistanceCall.objects.all())
    
    # Apply month filter
    if month:
        y, m = month.split('-')
        q_fire = [c for c in q_fire if c.form_date.year == int(y) and c.form_date.month == int(m)]
        q_general = [c for c in q_general if c.form_date.year == int(y) and c.form_date.month == int(m)]
        q_assist = [c for c in q_assist if c.form_date.year == int(y) and c.form_date.month == int(m)]
    
    # Apply start date filter
    if start:
        startdate = datetime.strptime(start, "%Y-%m-%d").date()
        q_fire = [c for c in q_fire if c.form_date >= startdate]
        q_general = [c for c in q_general if c.form_date >= startdate]
        q_assist = [c for c in q_assist if c.form_date >= startdate]
    
    # Apply end date filter
    if end:
        enddate = datetime.strptime(end, "%Y-%m-%d").date()
        q_fire = [c for c in q_fire if c.form_date <= enddate]
        q_general = [c for c in q_general if c.form_date <= enddate]
        q_assist = [c for c in q_assist if c.form_date <= enddate]
    
    # Calculate statistics using corrected logic
    total_cases = len(q_fire) + len(q_general) + len(q_assist)
    assistance_calls = len(q_assist)
    
    vehicle_accidents = sum([c.incident.lower() == "vehicle accident" for c in q_general])
    well_related = sum([c.incident.lower() == "well related" for c in q_general])
    lift_accidents = sum([c.incident.lower() == "lift accidents" for c in q_general])
    quarry = sum([c.incident.lower() == "quarry" for c in q_general])
    landslides = sum([c.incident.lower() == "landslides" for c in q_general])
    building_collapse = sum([c.incident.lower() == "building collapse" for c in q_general])
    tree_related = sum([c.incident.lower() == "tree related" for c in q_general])
    animal_calls = sum([c.incident.lower() == "animal calls" for c in q_general])

    animals_rescued = (sum([int(c.animals_rescued or 0) for c in q_fire]) +
                       sum([int(c.animals_rescued or 0) for c in q_general]))
    animals_lost = (sum([int(c.animals_lost or 0) for c in q_fire]) +
                    sum([int(c.animals_lost or 0) for c in q_general]))

    human_rescued = (
        sum([int(c.rescued_male or 0) for c in q_fire]) +
        sum([int(c.rescued_female or 0) for c in q_fire]) +
        sum([int(c.rescued_male or 0) for c in q_general]) +
        sum([int(c.rescued_female or 0) for c in q_general])
    )
    human_lost = (
        sum([int(c.deaths_male or 0) for c in q_fire]) +
        sum([int(c.deaths_female or 0) for c in q_fire]) +
        sum([int(c.deaths_male or 0) for c in q_general]) +
        sum([int(c.deaths_female or 0) for c in q_general])
    )
    
    # Water related incidents
    q_water = [c for c in q_fire if 'water' in c.incident.lower()]
    
    water_deaths_male = sum([int(c.deaths_male or 0) for c in q_water])
    water_deaths_female = sum([int(c.deaths_female or 0) for c in q_water])
    water_deaths_adult = sum([int(c.deaths_adult or 0) for c in q_water])
    water_deaths_child = sum([int(c.deaths_child or 0) for c in q_water])
    
    water_rescued_male = sum([int(c.rescued_male or 0) for c in q_water])
    water_rescued_female = sum([int(c.rescued_female or 0) for c in q_water])
    water_rescued_adult = sum([int(c.rescued_adult or 0) for c in q_water])
    water_rescued_child = sum([int(c.rescued_child or 0) for c in q_water])
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'incidents_analytics_{timestamp}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create CSV writer
    writer = csv.writer(response)
    
    # Write header with filter information
    writer.writerow(['Incidents Report Analytics'])
    writer.writerow(['Generated on:', datetime.now().strftime('%Y-%m-%d')])
    
    if month:
        writer.writerow(['Filtered by Month:', month])
    if start and end:
        writer.writerow(['Date Range:', f'{start} to {end}'])
    elif start:
        writer.writerow(['Start Date:', start])
    elif end:
        writer.writerow(['End Date:', end])
    
    writer.writerow([])  # Empty row for spacing
    
    # Write column headers
    writer.writerow(['Category', 'Total Cases', 'Notes'])
    
    # Write data rows
    writer.writerow(['Total Cases', total_cases, ''])
    writer.writerow(['Assistance Calls', assistance_calls, ''])
    writer.writerow(['Vehicle Accidents', vehicle_accidents, ''])
    writer.writerow(['Well Related', well_related, ''])
    writer.writerow(['Lift Accidents', lift_accidents, ''])
    writer.writerow(['Quarry', quarry, ''])
    writer.writerow(['Landslides', landslides, ''])
    writer.writerow(['Building Collapse', building_collapse, ''])
    writer.writerow(['Tree Related', tree_related, ''])
    writer.writerow(['Animal Calls', animal_calls, ''])
    writer.writerow(['Animal Rescued', animals_rescued, ''])
    writer.writerow(['Animal Lost', animals_lost, ''])
    writer.writerow(['Human Rescued', human_rescued, ''])
    writer.writerow(['Human Lost', human_lost, ''])
    
    # Water Related section
    writer.writerow([])
    writer.writerow(['Water Related Incidents', '', ''])
    writer.writerow(['Deaths - Male', water_deaths_male, ''])
    writer.writerow(['Deaths - Female', water_deaths_female, ''])
    writer.writerow(['Deaths - Adult (Above 18)', water_deaths_adult, ''])
    writer.writerow(['Deaths - Child (Below 18)', water_deaths_child, ''])
    writer.writerow(['Rescued - Male', water_rescued_male, ''])
    writer.writerow(['Rescued - Female', water_rescued_female, ''])
    writer.writerow(['Rescued - Adult (Above 18)', water_rescued_adult, ''])
    writer.writerow(['Rescued - Child (Below 18)', water_rescued_child, ''])
    
    return response